import pyodbc
import openpyxl
from setting import server,database_name

class writex_sql():

    def __init__(self,):
        self.conn = pyodbc.connect('DRIVER={SQL Server};SERVER=' + server + ';DATABASE=' + database_name + ';Trusted_Connection=yes')
# # Open the workbook and define the worksheet
    def run(self):
        self.cursor = self.conn.cursor()
        self.my_wb_obj = openpyxl.load_workbook("result.xlsx")
        sheet = self.my_wb_obj.active

        query1 = """
        CREATE TABLE reestr(
            "№ п/п" varchar(255),
            "Сокращенное наименование члена СРО" varchar(255),
            "ИНН" varchar(255),
            "Статус" varchar(255),
            "Тип" varchar(255),
            "Рег. Номер СРО" varchar(255),
            "Дата регистрации в реестре" varchar(255),
            "Дата прекращения членства в СРО" varchar(255),
            "Стоимость работ по одному договору подряда" varchar(255),
            "Размер обязательств по договорам подряда" varchar(255),
            "Сведение о приостановлении/возобновлении права: Дата" varchar(255),
            "Сведение о приостановлении/возобновлении права: Статус" varchar(255),
            "Архив: Номер свидетельства" varchar(255),
            "Архив: Дата выдачи" varchar(255),
            "Архив: Стоимость работ по одному ГП" varchar(255),
            "Архив: Статус действия свидетельства" varchar(255)
        )"""

        query = """
        INSERT INTO reestr (
            "№ п/п",
            "Сокращенное наименование члена СРО",
            "ИНН",
            "Статус",
            "Тип",
            "Рег. Номер СРО",
            "Дата регистрации в реестре",
            "Дата прекращения членства в СРО",
            "Стоимость работ по одному договору подряда",
            "Размер обязательств по договорам подряда",
            "Сведение о приостановлении/возобновлении права: Дата",
            "Сведение о приостановлении/возобновлении права: Статус",
            "Архив: Номер свидетельства",
            "Архив: Дата выдачи",
            "Архив: Стоимость работ по одному ГП",
            "Архив: Статус действия свидетельства"
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"""

        # execute create table
        try:
            self.cursor.execute(query1)
            self.conn.commit()
        except pyodbc.ProgrammingError:
            pass
        # grab existing row count in the database for validation later
        self.cursor.execute("TRUNCATE TABLE reestr")
        self.cursor.execute("SELECT count(*) FROM reestr")
        before_import = self.cursor.fetchone()

        for r in range(3, sheet.max_row):
            values = []
            for i in range(1,17):
                if sheet.cell(r,i).value==None:
                    values.append('')
                else:
                    values.append(sheet.cell(r,i).value)

            self.cursor.execute(query, tuple(values))

        self.conn.commit()
        self.conn.close()
        self.my_wb_obj.close()
if __name__ in '__main__':
    writex_sql().run()