import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

class Writexlsx():

    def __init__(self, baza):
        self.baza = baza

    def func_chunk(self,lst, n):
        for x in range(0, len(lst), n):
            e_c = lst[x: n + x]

            if len(e_c) < n:
                e_c = e_c + [None for y in range(n - len(e_c))]
            yield e_c

    def write(self,row,col,package):
        for j in range(len(package)):
            self.sheet.cell(row=row, column=j + col).font = self.font
            self.sheet.cell(row=row, column=j + col).alignment = self.alignment
            self.sheet.cell(row=row, column=j + col).value = package[j]

    def run(self):
        self.my_wb_obj = openpyxl.load_workbook("result.xlsx")
        self.font = Font(name='Times New Roman', size=12)
        self.alignment = Alignment(wrap_text=True)
        # workbook = xlsxwriter.Workbook('result.xlsx')
        for self.company in self.baza:
            for name, inn, status, tip, reg_number, date_reg, date_out, price_work, size, data, archiv in self.company:
                self.my_wb_obj.save("result.xlsx")
                self.sheet = self.my_wb_obj.active
                i = 3
                while self.sheet.cell(i, 1).value != None:
                    i += 1
                package_info = [i - 2, name, inn, status.strip(), tip, reg_number, date_reg, date_out,
                                price_work.replace("\n", '').replace("  ", '').strip(),
                                size.strip().replace("\n", '').replace("  ", '')]
                self.sheet.row_dimensions[i].height = 100
                self.write(i,1,package_info)
                data = list(self.func_chunk(data, 3))
                if len(data) == 1:
                    for k in data:
                        if None not in k:
                            package_data = [k[0].replace("\n", "").strip(), k[1].replace("\n", "").strip()]
                        else:
                            package_data = ["", ""]
                        self.write(i, 11, package_data)
                else:
                    i_a = i
                    for k in data:
                        package_data = [k[0].replace("\n", "").strip(), k[1].replace("\n", "").strip()]
                        self.write(i_a, 11, package_data)
                        i_a += 1
                    for n in range(len(data)):
                        package_info[0]=i-2
                        self.write(i, 1, package_info)
                        self.sheet.row_dimensions[i].height = 100
                        i += 1

                archiv = list(self.func_chunk(archiv, 8))
                if len(archiv) == 1:
                    for j in archiv:
                        package = [j[1].strip().replace(" &downarrow;", ""), j[2].strip(), j[4].replace("\n", "").strip(),
                                   j[5].strip().replace(" &downarrow;", "").replace("\n", "")]
                        self.write(i, 13, package)
                else:
                    i_a = i
                    for j in archiv:
                        package = [j[1].strip().replace(" &downarrow;", ""), j[2].strip(), j[4].replace("\n", "").strip(),
                                   j[5].strip().replace(" &downarrow;", "").replace("\n", "")]
                        self.write(i_a, 13, package)
                        i_a += 1
                    for n in range(len(archiv)):
                        package_info[0] = i-2
                        self.write(i, 1, package_info)
                        self.sheet.row_dimensions[i].height = 100
                        i += 1
        self.my_wb_obj.save('result.xlsx')
        self.my_wb_obj.close()
if __name__ == '__main__':
    date = [['ЗАО "ИСТ Казбек"', '2009002493',
             '\n\n                                                        Исключен\n                        ', 'ЮЛ',
             'СРО-С-147-24122009', '28.10.2014', '16.12.2020',
             '\n                                    не превышает шестьдесят миллионов рублей \n                    (Первый уровень ответственности)\n                            ',
             '\n                                    не превышает шестьдесят миллионов рублей \n                    (Первый уровень ответственности)\n                            ',
             ['\n                    16.12.2020\n                ',
              '\n                    Прекращение действия права\n                ',
              '\n                    Заявление о добровольном прекращении членства\n                '],
             ['1', '\n\n                    № 0178.01-2014-2009002493-C-147 &downarrow;\n                \n',
              '28.10.2014',
              'Решение Правления (протокол № 18 от 28.10.2014 г.)', '\n',
              '\n\n                    Действует\n                    &downarrow;\n                \n', 'Загрузка...',
              '\n                                    Загрузка...\n                            ']]]
    wr=Writexlsx(date)
    wr.run()
